"""Export lanes: CSV and a multi-sheet Excel workbook.

The workbook is the artefact people actually forward, print, and paste into
decks, so it is built to be *used* rather than merely produced:

  Summary    what this file contains, when it was produced, and the headline
             counts - so a forwarded copy is never context-free
  Filings    a native Excel Table: sortable, filterable, with live hyperlinks,
             real date cells, and conditional formatting on status
  By SRO     an SRO x status matrix with totals, the view people otherwise
             rebuild by hand as a pivot every single week
  Activity   period comparison, present only when one is supplied

Two details that matter more than they look:

  * dates are written as real ``date`` objects, not strings. A date-shaped
    string sorts lexically and breaks filtering, which is the single most
    common way an export becomes useless.
  * status colour is applied with conditional formatting rather than static
    fills, so it survives the user re-sorting the table.
"""

from __future__ import annotations

import csv
import datetime as dt
import sqlite3
from dataclasses import dataclass, field
from pathlib import Path
from typing import Sequence

from ..models import COLUMNS

# Excel refuses to open a workbook containing a cell longer than this.
_EXCEL_CELL_LIMIT = 32_000

# A leading =, +, - or @ makes Excel evaluate text as a formula. Summaries are
# third-party text from public websites, so they are neutralised before anyone
# double-clicks the file.
_FORMULA_PREFIXES = ("=", "+", "-", "@", "\t", "\r")

# House palette, shared with the dashboard and the email.
INK = "0B1F2A"
ACCENT = "0A66C2"
PAPER = "F5F7FA"
RULE = "D7DEE6"

STATUS_COLOURS: dict[str, str] = {
    "Filed": "06806F",
    "Approved": "1A7F37",
    "Immediately Effective": "0969DA",
    "Notice": "6639BA",
    "Amended": "8250DF",
    "Period Extended": "9A6700",
    "Proceedings Instituted": "BC4C00",
    "Suspended": "BC4C00",
    "Withdrawn": "6E7781",
    "Disapproved": "CF222E",
    "Unknown": "6E7781",
}

_HEADINGS = {
    "filing_no": "Filing No.",
    "sro": "SRO",
    "sro_family": "Family",
    "filing_year": "Year",
    "filing_date": "SEC Date",
    "status": "Status",
    "summary": "Summary",
    "release_number": "Release",
    "filing_url": "Document",
    "source": "Source",
    "source_url": "Listing",
    "notes": "Notes",
    "first_seen": "First Seen",
    "last_seen": "Last Seen",
    "seen_by": "Seen By",
}

_WIDTHS = {
    "filing_no": 22, "sro": 27, "sro_family": 13, "filing_year": 7,
    "filing_date": 12, "status": 22, "summary": 78, "release_number": 13,
    "filing_url": 30, "source": 24, "source_url": 24, "notes": 34,
    "first_seen": 18, "last_seen": 18, "seen_by": 26,
}


@dataclass(slots=True)
class PeriodComparison:
    """Two windows of filing activity, for the Activity sheet."""

    current_label: str
    previous_label: str
    current_total: int
    previous_total: int
    by_family: dict[str, tuple[int, int]] = field(default_factory=dict)
    by_status: dict[str, tuple[int, int]] = field(default_factory=dict)

    @property
    def delta(self) -> int:
        return self.current_total - self.previous_total


@dataclass(slots=True)
class ExportContext:
    """Everything the Summary sheet needs that the rows themselves cannot say."""

    scope: str = "All tracked filings"
    generated_at: dt.datetime | None = None
    source_health: Sequence[sqlite3.Row] = ()
    comparison: PeriodComparison | None = None
    total_tracked: int | None = None


# ---------------------------------------------------------------------------
# Shared helpers
# ---------------------------------------------------------------------------


def _rows(records: Sequence[sqlite3.Row]) -> list[dict[str, object]]:
    out: list[dict[str, object]] = []
    for record in records:
        keys = record.keys() if hasattr(record, "keys") else []
        out.append({c: (record[c] if c in keys else "") for c in COLUMNS})
    return out


def _safe(value: object, *, limit: int | None = None) -> object:
    if not isinstance(value, str):
        return value
    text = value
    if text.startswith(_FORMULA_PREFIXES):
        text = "'" + text
    if limit and len(text) > limit:
        text = text[: limit - 1] + "…"
    return text


def timestamped(directory: Path, stem: str, suffix: str) -> Path:
    stamp = dt.datetime.now().strftime("%Y%m%d-%H%M%S")
    directory.mkdir(parents=True, exist_ok=True)
    return directory / f"{stem}-{stamp}{suffix}"


def to_csv(records: Sequence[sqlite3.Row], path: Path) -> Path:
    """utf-8-sig so Excel opens it with the right encoding on a double-click."""
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", newline="", encoding="utf-8-sig") as handle:
        writer = csv.DictWriter(handle, fieldnames=list(COLUMNS))
        writer.writeheader()
        for row in _rows(records):
            writer.writerow({k: _safe(v) for k, v in row.items()})
    return path


# ---------------------------------------------------------------------------
# Excel
# ---------------------------------------------------------------------------


def to_excel(
    records: Sequence[sqlite3.Row],
    path: Path,
    *,
    title: str = "Rule Filings",
    context: ExportContext | None = None,
) -> Path:
    from openpyxl import Workbook

    context = context or ExportContext()
    context.generated_at = context.generated_at or dt.datetime.now()

    path.parent.mkdir(parents=True, exist_ok=True)
    rows = _rows(records)

    book = Workbook()
    book.remove(book.active)

    _summary_sheet(book, rows, context, title)
    _filings_sheet(book, rows)
    _matrix_sheet(book, rows)
    if context.comparison:
        _activity_sheet(book, context.comparison)

    book.properties.title = f"SRO Rule Filings - {title}"
    book.properties.creator = "SRO Filing Tracker"
    book.properties.created = context.generated_at
    book.properties.description = (
        f"{len(rows):,} SRO rule filings. Scope: {context.scope}."
    )

    book.active = 0
    book.save(path)
    return path


# ---- sheet builders -------------------------------------------------------


def _style_title(cell, size: int = 16) -> None:
    from openpyxl.styles import Font

    cell.font = Font(name="Segoe UI", size=size, bold=True, color=INK)


def _label(sheet, row: int, text: str, value: object, *, bold: bool = False) -> None:
    from openpyxl.styles import Font

    key = sheet.cell(row=row, column=2, value=text)
    key.font = Font(name="Segoe UI", size=10, color="5A6B7A")
    val = sheet.cell(row=row, column=3, value=value)
    val.font = Font(name="Segoe UI", size=10, bold=bold, color=INK)


def _summary_sheet(book, rows, context: ExportContext, title: str) -> None:
    from openpyxl.styles import Alignment, Font, PatternFill

    sheet = book.create_sheet("Summary")
    sheet.sheet_view.showGridLines = False
    for column, width in (("A", 2), ("B", 26), ("C", 34), ("D", 3),
                          ("E", 26), ("F", 12), ("G", 12)):
        sheet.column_dimensions[column].width = width

    banner = sheet.cell(row=2, column=2, value="SRO RULE FILINGS")
    _style_title(banner, 18)
    sheet.cell(row=3, column=2, value=title).font = Font(
        name="Segoe UI", size=11, color="5A6B7A")

    generated = context.generated_at or dt.datetime.now()
    _label(sheet, 5, "Generated", generated.strftime("%d %B %Y, %H:%M"))
    _label(sheet, 6, "Scope", context.scope)
    _label(sheet, 7, "Rows in this file", len(rows), bold=True)
    if context.total_tracked is not None:
        _label(sheet, 8, "Total tracked", context.total_tracked)

    dates = [r["filing_date"] for r in rows if r["filing_date"]]
    if dates:
        _label(sheet, 9, "Date range", f"{min(dates)}  to  {max(dates)}")

    def block(start_row: int, heading: str, counts: dict[str, int],
              colour_key: bool = False) -> int:
        head = sheet.cell(row=start_row, column=2, value=heading)
        head.font = Font(name="Segoe UI", size=11, bold=True, color=INK)
        head.fill = PatternFill("solid", fgColor=PAPER)
        sheet.cell(row=start_row, column=3, value="").fill = PatternFill("solid", fgColor=PAPER)
        line = start_row + 1
        for name, count in sorted(counts.items(), key=lambda kv: (-kv[1], kv[0])):
            key = sheet.cell(row=line, column=2, value=name)
            key.font = Font(
                name="Segoe UI", size=10,
                color=STATUS_COLOURS.get(name, INK) if colour_key else INK,
                bold=colour_key and name in STATUS_COLOURS,
            )
            val = sheet.cell(row=line, column=3, value=count)
            val.font = Font(name="Segoe UI", size=10, color=INK)
            val.alignment = Alignment(horizontal="left")
            line += 1
        return line + 1

    families: dict[str, int] = {}
    statuses: dict[str, int] = {}
    sros: dict[str, int] = {}
    for row in rows:
        families[str(row["sro_family"] or "—")] = families.get(str(row["sro_family"] or "—"), 0) + 1
        statuses[str(row["status"] or "—")] = statuses.get(str(row["status"] or "—"), 0) + 1
        sros[str(row["sro"] or "—")] = sros.get(str(row["sro"] or "—"), 0) + 1

    next_row = block(11, "BY FAMILY", families)
    next_row = block(next_row, "BY STATUS", statuses, colour_key=True)

    # Source health, in a second column so the sheet reads on one screen.
    if context.source_health:
        head = sheet.cell(row=11, column=5, value="SOURCE HEALTH")
        head.font = Font(name="Segoe UI", size=11, bold=True, color=INK)
        head.fill = PatternFill("solid", fgColor=PAPER)
        for offset, entry in enumerate(context.source_health, start=12):
            keys = entry.keys() if hasattr(entry, "keys") else []
            name = entry["source"] if "source" in keys else ""
            state = str(entry["status"]) if "status" in keys else ""
            count = entry["records"] if "records" in keys else ""
            sheet.cell(row=offset, column=5, value=name).font = Font(
                name="Consolas", size=9, color=INK)
            cell = sheet.cell(row=offset, column=6, value=state)
            cell.font = Font(
                name="Segoe UI", size=9, bold=True,
                color={"ok": "1A7F37", "degraded": "9A6700"}.get(state, "CF222E"),
            )
            sheet.cell(row=offset, column=7, value=count).font = Font(
                name="Segoe UI", size=9, color="5A6B7A")

    foot = sheet.cell(
        row=max(next_row, 14) + 1, column=2,
        value="Source: SEC Self-Regulatory Organization Rulemaking listings and "
              "exchange rule-filing feeds.",
    )
    foot.font = Font(name="Segoe UI", size=9, italic=True, color="8A97A3")


def _filings_sheet(book, rows) -> None:
    from openpyxl.formatting.rule import CellIsRule
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.table import Table, TableStyleInfo

    sheet = book.create_sheet("Filings")
    sheet.sheet_view.showGridLines = False

    header_font = Font(name="Segoe UI", size=10, bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor=INK)
    for index, column in enumerate(COLUMNS, start=1):
        cell = sheet.cell(row=1, column=index, value=_HEADINGS.get(column, column))
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(vertical="center", horizontal="left")
    sheet.row_dimensions[1].height = 22

    body = Font(name="Segoe UI", size=10)
    link_font = Font(name="Segoe UI", size=10, color=ACCENT, underline="single")
    mono = Font(name="Consolas", size=10)
    wrap = Alignment(vertical="top", wrap_text=True)
    top = Alignment(vertical="top")

    for row_index, row in enumerate(rows, start=2):
        for col_index, column in enumerate(COLUMNS, start=1):
            raw = row[column]
            cell = sheet.cell(row=row_index, column=col_index)

            if column == "filing_date" and raw:
                # A real date, so sorting and date filters behave.
                try:
                    cell.value = dt.date.fromisoformat(str(raw)[:10])
                    cell.number_format = "yyyy-mm-dd"
                except ValueError:
                    cell.value = _safe(raw)
                cell.font = body
                cell.alignment = top
                continue

            if column == "filing_no":
                cell.value = _safe(raw)
                url = row.get("filing_url")
                if url and str(url).startswith("http"):
                    # Hyperlink on the identifier, not on a raw URL column -
                    # that is where people click.
                    cell.hyperlink = str(url)
                    cell.font = Font(name="Consolas", size=10, color=ACCENT,
                                     underline="single", bold=True)
                else:
                    cell.font = Font(name="Consolas", size=10, bold=True)
                cell.alignment = top
                continue

            if column in {"filing_url", "source_url"} and raw and str(raw).startswith("http"):
                cell.value = "Open"
                cell.hyperlink = str(raw)
                cell.font = link_font
                cell.alignment = top
                continue

            if column == "filing_year" and raw:
                try:
                    cell.value = int(raw)
                except (TypeError, ValueError):
                    cell.value = _safe(raw)
                cell.font = body
                cell.alignment = top
                continue

            cell.value = _safe(raw, limit=_EXCEL_CELL_LIMIT)
            cell.font = mono if column == "release_number" else body
            cell.alignment = wrap if column in {"summary", "notes"} else top

    last_row = max(len(rows) + 1, 2)
    last_col = get_column_letter(len(COLUMNS))

    for index, column in enumerate(COLUMNS, start=1):
        sheet.column_dimensions[get_column_letter(index)].width = _WIDTHS.get(column, 18)

    # A native Excel Table: users get sort and filter controls without touching
    # the ribbon. Banding comes from the table style; status colour is applied
    # by conditional formatting so it survives re-sorting.
    if rows:
        table = Table(displayName="Filings", ref=f"A1:{last_col}{last_row}")
        table.tableStyleInfo = TableStyleInfo(
            name="TableStyleLight8", showRowStripes=True, showColumnStripes=False,
            showFirstColumn=False, showLastColumn=False,
        )
        sheet.add_table(table)

        status_col = get_column_letter(COLUMNS.index("status") + 1)
        status_range = f"{status_col}2:{status_col}{last_row}"
        for status, colour in STATUS_COLOURS.items():
            sheet.conditional_formatting.add(
                status_range,
                CellIsRule(operator="equal", formula=[f'"{status}"'],
                           font=Font(name="Segoe UI", size=10, bold=True, color=colour)),
            )

    # Freeze at B2: the header row and the filing number both stay visible when
    # scrolling right through the summary.
    sheet.freeze_panes = "B2"

    sheet.page_setup.orientation = "landscape"
    sheet.page_setup.fitToWidth = 1
    sheet.page_setup.fitToHeight = 0
    sheet.sheet_properties.pageSetUpPr.fitToPage = True
    sheet.print_title_rows = "1:1"
    sheet.oddFooter.right.text = "Page &P of &N"
    sheet.oddFooter.left.text = "SRO Rule Filings"


def _matrix_sheet(book, rows) -> None:
    """SRO x status matrix - the pivot people otherwise rebuild by hand."""
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    sheet = book.create_sheet("By SRO")
    sheet.sheet_view.showGridLines = False

    statuses: list[str] = []
    for status in STATUS_COLOURS:
        if any(str(r["status"]) == status for r in rows):
            statuses.append(status)
    for row in rows:
        value = str(row["status"] or "—")
        if value not in statuses:
            statuses.append(value)

    matrix: dict[tuple[str, str], int] = {}
    families: dict[str, str] = {}
    for row in rows:
        sro = str(row["sro"] or "—")
        families[sro] = str(row["sro_family"] or "—")
        key = (sro, str(row["status"] or "—"))
        matrix[key] = matrix.get(key, 0) + 1

    header_font = Font(name="Segoe UI", size=10, bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor=INK)
    headers = ["SRO", "Family", *statuses, "Total"]
    for index, name in enumerate(headers, start=1):
        cell = sheet.cell(row=1, column=index, value=name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(vertical="center", wrap_text=True, horizontal="center"
                                   if index > 2 else "left")
    sheet.row_dimensions[1].height = 34

    ordered = sorted(families, key=lambda s: (families[s], s))
    for offset, sro in enumerate(ordered, start=2):
        sheet.cell(row=offset, column=1, value=sro).font = Font(name="Segoe UI", size=10)
        sheet.cell(row=offset, column=2, value=families[sro]).font = Font(
            name="Segoe UI", size=10, color="5A6B7A")
        total = 0
        for index, status in enumerate(statuses, start=3):
            count = matrix.get((sro, status), 0)
            total += count
            cell = sheet.cell(row=offset, column=index, value=count or None)
            cell.font = Font(name="Segoe UI", size=10,
                             color=STATUS_COLOURS.get(status, INK))
            cell.alignment = Alignment(horizontal="center")
        cell = sheet.cell(row=offset, column=len(headers), value=total)
        cell.font = Font(name="Segoe UI", size=10, bold=True, color=INK)
        cell.alignment = Alignment(horizontal="center")

    total_row = len(ordered) + 2
    label = sheet.cell(row=total_row, column=1, value="TOTAL")
    label.font = Font(name="Segoe UI", size=10, bold=True, color=INK)
    label.fill = PatternFill("solid", fgColor=PAPER)
    sheet.cell(row=total_row, column=2, value="").fill = PatternFill("solid", fgColor=PAPER)
    for index, status in enumerate(statuses, start=3):
        cell = sheet.cell(row=total_row, column=index,
                          value=sum(matrix.get((s, status), 0) for s in ordered) or None)
        cell.font = Font(name="Segoe UI", size=10, bold=True, color=INK)
        cell.fill = PatternFill("solid", fgColor=PAPER)
        cell.alignment = Alignment(horizontal="center")
    cell = sheet.cell(row=total_row, column=len(headers), value=len(rows))
    cell.font = Font(name="Segoe UI", size=10, bold=True, color=INK)
    cell.fill = PatternFill("solid", fgColor=PAPER)
    cell.alignment = Alignment(horizontal="center")

    sheet.column_dimensions["A"].width = 30
    sheet.column_dimensions["B"].width = 14
    for index in range(3, len(headers) + 1):
        sheet.column_dimensions[get_column_letter(index)].width = 13
    sheet.freeze_panes = "C2"


def _activity_sheet(book, comparison: PeriodComparison) -> None:
    from openpyxl.styles import Alignment, Font, PatternFill

    sheet = book.create_sheet("Activity")
    sheet.sheet_view.showGridLines = False
    for column, width in (("A", 2), ("B", 30), ("C", 16), ("D", 16), ("E", 12)):
        sheet.column_dimensions[column].width = width

    _style_title(sheet.cell(row=2, column=2, value="PERIOD COMPARISON"), 14)

    header_font = Font(name="Segoe UI", size=10, bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor=INK)

    def header(row: int, first: str) -> None:
        for index, name in enumerate(
                [first, comparison.current_label, comparison.previous_label, "Change"], start=2):
            cell = sheet.cell(row=row, column=index, value=name)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center" if index > 2 else "left",
                                       vertical="center", wrap_text=True)
        sheet.row_dimensions[row].height = 30

    def line(row: int, name: str, current: int, previous: int, *, bold: bool = False) -> None:
        sheet.cell(row=row, column=2, value=name).font = Font(
            name="Segoe UI", size=10, bold=bold, color=INK)
        for index, value in ((3, current), (4, previous)):
            cell = sheet.cell(row=row, column=index, value=value)
            cell.font = Font(name="Segoe UI", size=10, bold=bold, color=INK)
            cell.alignment = Alignment(horizontal="center")
        delta = current - previous
        cell = sheet.cell(row=row, column=5,
                          value=("+" if delta > 0 else "") + str(delta) if delta else "—")
        cell.font = Font(name="Segoe UI", size=10, bold=True,
                         color="1A7F37" if delta > 0 else ("CF222E" if delta < 0 else "8A97A3"))
        cell.alignment = Alignment(horizontal="center")

    header(4, "Total filings")
    line(5, "All SROs", comparison.current_total, comparison.previous_total, bold=True)

    row = 7
    if comparison.by_family:
        header(row, "By family")
        row += 1
        for name, (current, previous) in sorted(
                comparison.by_family.items(), key=lambda kv: -kv[1][0]):
            line(row, name, current, previous)
            row += 1
        row += 1

    if comparison.by_status:
        header(row, "By status")
        row += 1
        for name, (current, previous) in sorted(
                comparison.by_status.items(), key=lambda kv: -kv[1][0]):
            line(row, name, current, previous)
            row += 1

    note = sheet.cell(
        row=row + 1, column=2,
        value="Counts are by SEC issue date, so a re-run or a missed day never "
              "changes a historical period.",
    )
    note.font = Font(name="Segoe UI", size=9, italic=True, color="8A97A3")
