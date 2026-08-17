"""Weekly comparison, and the workbook.

The comparison logic is the part most likely to be quietly wrong: it is easy to
write something that looks right on a full dataset and reports nonsense after a
backfill or a missed run. These tests pin the semantics.
"""

from __future__ import annotations

import datetime as dt

import pytest
from openpyxl import load_workbook

from sro_tracker import exports, report as report_module
from sro_tracker.models import (
    STATUS_APPROVED,
    STATUS_FILED,
    STATUS_IMMEDIATELY_EFFECTIVE,
    STATUS_NOTICE,
    STATUS_UNKNOWN,
    Filing,
)
from sro_tracker.sources.exchange import parse_rss

TODAY = dt.date(2026, 8, 17)
END = TODAY + dt.timedelta(days=1)          # exclusive
THIS_WEEK = END - dt.timedelta(days=7)      # 11 Aug
LAST_WEEK = THIS_WEEK - dt.timedelta(days=7)  # 04 Aug


def make(no: str, *, date: dt.date, status=STATUS_NOTICE, family="Cboe",
         sro="Cboe BZX Exchange", summary="a proposed rule change") -> Filing:
    return Filing.build(
        filing_no=no, sro=sro, sro_family=family, summary=summary,
        status=status, filing_date=date, source="SEC",
        filing_url=f"https://example.test/{no}.pdf",
    )


@pytest.fixture
def seeded(tmp_store):
    """Three this week, two last week, one long before."""
    tmp_store.commit_filings(tmp_store.start_run(), [
        make("SR-CboeBZX-2026-1", date=dt.date(2026, 8, 14), status=STATUS_APPROVED),
        make("SR-CboeBZX-2026-2", date=dt.date(2026, 8, 12)),
        make("SR-NYSE-2026-3", date=dt.date(2026, 8, 11), family="NYSE",
             sro="New York Stock Exchange"),
        make("SR-CboeBZX-2026-4", date=dt.date(2026, 8, 7)),
        make("SR-CboeBZX-2026-5", date=dt.date(2026, 8, 5)),
        make("SR-CboeBZX-2026-6", date=dt.date(2026, 1, 5)),
    ])
    return tmp_store


# ---------------------------------------------------------------------------
# Period comparison
# ---------------------------------------------------------------------------


def test_periods_split_correctly(seeded):
    built = report_module.build(seeded, days=7, end=END)
    assert built.current.total == 3
    assert built.previous.total == 2
    assert built.delta == 1


def test_periods_are_measured_by_filing_date_not_scrape_time(seeded):
    """Everything above was committed just now, in one run. If the comparison
    keyed off the change log, both periods would read as this week's news and
    week-over-week would be meaningless after any backfill."""
    built = report_module.build(seeded, days=7, end=END)
    assert built.previous.total == 2, "prior period must not collapse to zero"
    assert seeded.count() == 6


def test_period_boundaries_are_half_open(seeded):
    """[start, end) - a filing dated exactly on the boundary belongs to exactly
    one period, never both and never neither."""
    built = report_module.build(seeded, days=7, end=END)
    current = {r["filing_no"] for r in built.current.filings}
    assert "SR-NYSE-2026-3" in current       # dated 11 Aug == current start
    assert "SR-CboeBZX-2026-4" not in current  # dated 07 Aug -> prior period
    assert built.current.total + built.previous.total == 5


def test_breakdowns_cover_both_periods(seeded):
    built = report_module.build(seeded, days=7, end=END)
    families = dict((n, (c, p)) for n, c, p in built.families)
    assert families["NYSE"] == (1, 0)
    assert families["Cboe"] == (2, 2)


def test_subject_reports_the_delta(seeded):
    built = report_module.build(seeded, days=7, end=END)
    assert "3 new" in built.subject
    assert "+1" in built.subject


def test_quiet_period_is_stated_plainly(tmp_store):
    built = report_module.build(tmp_store, days=7, end=END)
    assert built.is_quiet
    assert "no activity" in built.subject.lower()
    assert "No filings were issued" in report_module.render_text(built)


def test_status_change_on_an_older_filing_is_reported(tmp_store):
    """A 2025 filing approved this morning is this week's news even though its
    date is old - the one thing that legitimately keys off observation time."""
    tmp_store.commit_filings(tmp_store.start_run(), [
        make("SR-CboeBZX-2025-9", date=dt.date(2025, 3, 3), status=STATUS_NOTICE)])
    tmp_store.commit_filings(tmp_store.start_run(), [
        make("SR-CboeBZX-2025-9", date=dt.date(2025, 3, 3), status=STATUS_APPROVED)])

    built = report_module.build(tmp_store, days=7, end=END)
    assert len(built.status_changes) == 1
    _row, before, after = built.status_changes[0]
    assert (before, after) == (STATUS_NOTICE, STATUS_APPROVED)


def test_new_filings_are_not_also_listed_as_status_changes(seeded):
    seeded.commit_filings(seeded.start_run(), [
        make("SR-CboeBZX-2026-2", date=dt.date(2026, 8, 12), status=STATUS_APPROVED)])
    built = report_module.build(seeded, days=7, end=END)
    listed = {r["filing_no"] for r in built.current.filings}
    changed = {r["filing_no"] for r, _b, _a in built.status_changes}
    assert not (listed & changed), "a filing must not appear in both sections"


# ---------------------------------------------------------------------------
# Email rendering
# ---------------------------------------------------------------------------


def test_html_is_outlook_safe(seeded):
    """Desktop Outlook renders with Word: no flexbox, grid, float, or <style>."""
    html = report_module.render_html(report_module.build(seeded, days=7, end=END))
    lowered = html.lower()
    for banned in ("display:flex", "display: flex", "display:grid", "display: grid",
                   "<style", "float:", "position:absolute"):
        assert banned not in lowered, f"{banned!r} does not survive Word rendering"
    assert "<table" in lowered
    assert 'bgcolor=' in lowered, "background colours need the attribute for Word"


def test_html_escapes_filing_text(tmp_store):
    tmp_store.commit_filings(tmp_store.start_run(), [
        make("SR-CboeBZX-2026-1", date=dt.date(2026, 8, 14),
             summary='<script>alert("x")</script> & more')])
    html = report_module.render_html(report_module.build(tmp_store, days=7, end=END))
    assert "<script>" not in html
    assert "&lt;script&gt;" in html


def test_long_filing_lists_are_capped_with_a_pointer(tmp_store):
    tmp_store.commit_filings(tmp_store.start_run(), [
        make(f"SR-CboeBZX-2026-{i}", date=dt.date(2026, 8, 14)) for i in range(1, 41)])
    built = report_module.build(tmp_store, days=7, end=END)
    html = report_module.render_html(built, max_rows=10)
    assert "further filing" in html, "truncation must be disclosed, not silent"


def test_text_alternative_covers_the_same_ground(seeded):
    text = report_module.render_text(report_module.build(seeded, days=7, end=END))
    assert "BY FAMILY" in text and "NEW FILINGS" in text
    assert "SR-CboeBZX-2026-1" in text


# ---------------------------------------------------------------------------
# Workbook
# ---------------------------------------------------------------------------


def test_workbook_has_every_sheet(seeded, tmp_path):
    built = report_module.build(seeded, days=7, end=END)
    context = exports.ExportContext(
        scope="test", comparison=report_module.comparison_for_export(built),
        total_tracked=seeded.count())
    path = exports.to_excel(seeded.query(), tmp_path / "wb.xlsx", context=context)

    book = load_workbook(path)
    assert book.sheetnames == ["Summary", "Filings", "By SRO", "Activity"]


def test_dates_are_real_dates_not_strings(seeded, tmp_path):
    """A date-shaped string sorts lexically and breaks Excel's date filters."""
    path = exports.to_excel(seeded.query(), tmp_path / "wb.xlsx")
    sheet = load_workbook(path)["Filings"]
    column = exports.COLUMNS.index("filing_date") + 1
    value = sheet.cell(row=2, column=column).value
    assert isinstance(value, (dt.date, dt.datetime))


def test_filing_number_carries_the_hyperlink(seeded, tmp_path):
    path = exports.to_excel(seeded.query(), tmp_path / "wb.xlsx")
    cell = load_workbook(path)["Filings"].cell(row=2, column=1)
    assert cell.hyperlink is not None
    assert cell.hyperlink.target.startswith("https://")


def test_filings_sheet_is_a_native_table(seeded, tmp_path):
    path = exports.to_excel(seeded.query(), tmp_path / "wb.xlsx")
    sheet = load_workbook(path)["Filings"]
    assert "Filings" in sheet.tables
    assert sheet.freeze_panes == "B2"


def test_formula_injection_is_neutralised(tmp_store, tmp_path):
    tmp_store.commit_filings(tmp_store.start_run(), [
        make("SR-CboeBZX-2026-1", date=dt.date(2026, 8, 14),
             summary='=cmd|" /c calc"!A0')])
    rows = tmp_store.query()

    sheet = load_workbook(exports.to_excel(rows, tmp_path / "wb.xlsx"))["Filings"]
    column = exports.COLUMNS.index("summary") + 1
    assert not str(sheet.cell(row=2, column=column).value).startswith("=")

    csv_text = (exports.to_csv(rows, tmp_path / "out.csv")).read_text(encoding="utf-8-sig")
    assert "\n=cmd" not in csv_text and not csv_text.split("\n")[1].startswith("=")


def test_empty_result_still_produces_a_valid_workbook(tmp_store, tmp_path):
    """An export of nothing must open, not corrupt."""
    path = exports.to_excel(tmp_store.query(), tmp_path / "empty.xlsx")
    book = load_workbook(path)
    assert "Filings" in book.sheetnames
    assert book["Summary"]["C7"].value == 0


def test_matrix_totals_reconcile(seeded, tmp_path):
    rows = seeded.query()
    path = exports.to_excel(rows, tmp_path / "wb.xlsx")
    sheet = load_workbook(path)["By SRO"]
    grand = None
    for row in sheet.iter_rows():
        if row[0].value == "TOTAL":
            grand = row[-1].value
    assert grand == len(rows), "matrix total must equal the row count"


# ---------------------------------------------------------------------------
# Filed status
# ---------------------------------------------------------------------------


def test_feed_only_filings_are_marked_filed():
    """Not "Unknown": we know the exchange published it and the SEC has not
    acted yet."""
    xml = """<?xml version="1.0"?><rss><channel>
      <item><title>SR-CboeBZX-2026-066</title>
      <description>The Exchange proposes to amend its Fees Schedule.</description>
      <pubDate>Thu, 13 Aug 2026 00:00:00 -0400</pubDate></item>
    </channel></rss>"""
    filings = parse_rss(xml, source_url="u", source_label="Cboe BZX feed")
    assert filings[0].status == STATUS_FILED


def test_feed_status_is_used_when_the_description_states_one():
    xml = """<?xml version="1.0"?><rss><channel>
      <item><title>SR-CboeBZX-2026-070</title>
      <description>Order Approving a Proposed Rule Change</description></item>
    </channel></rss>"""
    assert parse_rss(xml, source_url="u", source_label="l")[0].status == STATUS_APPROVED


def test_sec_status_supersedes_filed():
    """The edge label must never survive once the SEC release exists."""
    from sro_tracker.pipeline import reconcile
    from sro_tracker.sources import STATUS_OK, TIER_EDGE, TIER_SPINE, SourceResult

    edge = Filing.build(filing_no="SR-CboeBZX-2026-1", sro="Cboe BZX Exchange",
                        sro_family="Cboe", summary="x", status=STATUS_FILED,
                        source="feed")
    spine = Filing.build(filing_no="SR-CboeBZX-2026-1", sro="Cboe BZX Exchange",
                         sro_family="Cboe", summary="x",
                         status=STATUS_IMMEDIATELY_EFFECTIVE, source="SEC")
    merged = reconcile([
        SourceResult("SEC:x", TIER_SPINE, STATUS_OK, [spine]),
        SourceResult("feed:x", TIER_EDGE, STATUS_OK, [edge]),
    ])
    assert merged[0].status == STATUS_IMMEDIATELY_EFFECTIVE


def test_filed_outranks_unknown():
    from sro_tracker.models import STATUS_RANK

    assert STATUS_RANK[STATUS_FILED] > STATUS_RANK[STATUS_UNKNOWN]
    assert STATUS_RANK[STATUS_FILED] < STATUS_RANK[STATUS_NOTICE]
