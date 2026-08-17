"""Hyperlink integrity.

Every filing must reach its own PDF, and every SRO must reach its own site.
A broken or mislabelled link is worse than a missing one: it looks authoritative
and sends the reader somewhere wrong.

The live-reachability check is marked ``network`` and skipped by default, so the
normal suite stays offline. Run it deliberately:

    pytest -m network
"""

from __future__ import annotations

import datetime as dt

import pytest
from openpyxl import load_workbook

from sro_tracker import exports, registry, report as report_module
from sro_tracker.models import Filing
from sro_tracker.sources.sec_sro import parse_listing


# ---------------------------------------------------------------------------
# Registry coverage
# ---------------------------------------------------------------------------


def test_every_sro_has_a_website():
    missing = [s.key for s in registry.all_sros() if not s.website]
    assert not missing, f"no website registered for: {missing}"


def test_every_website_is_https():
    for sro in registry.all_sros():
        assert sro.website.startswith("https://"), f"{sro.key}: {sro.website}"


def test_website_lookup_resolves_by_display_name():
    """Records store the display name, so that is the lookup that must work."""
    for sro in registry.all_sros():
        assert registry.website_for(sro.name) == sro.website
    assert registry.website_for("New York Stock Exchange").startswith("https://www.nyse.com")
    assert registry.website_for("FINRA").startswith("https://www.finra.org")


def test_website_lookup_is_forgiving_of_case_and_padding():
    assert registry.website_for("  finra  ") == registry.get("finra").website


def test_unknown_sro_name_returns_empty_not_a_wrong_link():
    assert registry.website_for("Not A Real Exchange") == ""
    assert registry.website_for("") == ""


def test_each_family_points_at_its_own_operator():
    """Guards against a copy-paste that sends one family to another's site."""
    expected = {
        "NYSE": "nyse.com",
        "Nasdaq": "nasdaq.com",
        "Cboe": "cboe.com",
        "FINRA": "finra.org",
    }
    for sro in registry.all_sros():
        host = expected.get(sro.family)
        if host:
            assert host in sro.website, f"{sro.key} -> {sro.website}"


# ---------------------------------------------------------------------------
# Filing PDF links
# ---------------------------------------------------------------------------


def test_parsed_filings_all_carry_a_pdf(sec_listing_html):
    filings = parse_listing(sec_listing_html, sro=registry.get("nyse"),
                            source_url="https://example.test/x")
    for filing in filings:
        assert filing.filing_url, f"{filing.filing_no} has no document link"
        assert filing.filing_url.startswith("https://www.sec.gov/")
        assert filing.filing_url.lower().endswith(".pdf")


def test_pdf_link_is_the_release_not_an_exhibit(sec_listing_html):
    """The release PDF is the filing; -ex5 is an attachment. Linking the
    attachment sends the reader to a fragment of the document."""
    filings = parse_listing(sec_listing_html, sro=registry.get("nyse"),
                            source_url="https://example.test/x")
    for filing in filings:
        assert "-ex" not in filing.filing_url.lower(), filing.filing_url


def test_pdf_path_matches_the_release_number(sec_listing_html):
    """The SEC names the file after the release, so a mismatch means rows and
    links have been crossed."""
    filings = parse_listing(sec_listing_html, sro=registry.get("nyse"),
                            source_url="https://example.test/x")
    checked = 0
    for filing in filings:
        if filing.release_number:
            assert filing.release_number in filing.filing_url, (
                f"{filing.filing_no}: release {filing.release_number} "
                f"not in {filing.filing_url}")
            checked += 1
    assert checked >= 5


# ---------------------------------------------------------------------------
# Links in the rendered surfaces
# ---------------------------------------------------------------------------


def _seed(store):
    store.commit_filings(store.start_run(), [
        Filing.build(
            filing_no="SR-NYSE-2026-38", sro="New York Stock Exchange",
            sro_family="NYSE", summary="Amend the price list", status="Notice",
            filing_date=dt.date(2026, 8, 14), release_number="34-106138",
            filing_url="https://www.sec.gov/files/rules/sro/nyse/2026/34-106138.pdf",
            source="SEC"),
    ])


def test_email_links_the_filing_to_its_pdf_and_the_sro_to_its_site(tmp_store):
    _seed(tmp_store)
    built = report_module.build(tmp_store, days=7,
                                end=dt.date(2026, 8, 18))
    html = report_module.render_html(built)
    assert "https://www.sec.gov/files/rules/sro/nyse/2026/34-106138.pdf" in html
    assert registry.get("nyse").website in html


def test_workbook_links_the_sro_cell(tmp_store, tmp_path):
    _seed(tmp_store)
    path = exports.to_excel(tmp_store.query(), tmp_path / "wb.xlsx")
    sheet = load_workbook(path)["Filings"]
    sro_cell = sheet.cell(row=2, column=exports.COLUMNS.index("sro") + 1)
    assert sro_cell.hyperlink is not None
    assert sro_cell.hyperlink.target == registry.get("nyse").website


def test_workbook_links_the_filing_number_to_the_pdf(tmp_store, tmp_path):
    _seed(tmp_store)
    path = exports.to_excel(tmp_store.query(), tmp_path / "wb.xlsx")
    cell = load_workbook(path)["Filings"].cell(row=2, column=1)
    assert cell.hyperlink.target.endswith("34-106138.pdf")


# ---------------------------------------------------------------------------
# Live reachability - opt in with `pytest -m network`
# ---------------------------------------------------------------------------


@pytest.mark.network
def test_every_registered_website_resolves():
    """A 403 counts as reachable: some sites refuse automated clients but serve
    a human browser, which is all a link has to do. Only 404 or DNS failure is
    a broken link."""
    import urllib.error
    import urllib.request

    ua = ("Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
          "(KHTML, like Gecko) Chrome/125.0 Safari/537.36")
    broken: list[str] = []
    for sro in registry.all_sros():
        request = urllib.request.Request(sro.website, headers={"User-Agent": ua})
        try:
            with urllib.request.urlopen(request, timeout=25) as response:
                if response.status >= 400:
                    broken.append(f"{sro.key} -> {response.status}")
        except urllib.error.HTTPError as exc:
            if exc.code not in (403, 405, 406, 429):
                broken.append(f"{sro.key} -> HTTP {exc.code}")
        except Exception as exc:  # noqa: BLE001
            broken.append(f"{sro.key} -> {type(exc).__name__}")
    assert not broken, f"unreachable SRO websites: {broken}"
